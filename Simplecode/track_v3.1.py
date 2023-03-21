import warnings
import matplotlib.pyplot as plt
from scipy.spatial import distance
import numpy as np
from scipy.optimize import linear_sum_assignment
import math
import openpyxl
import os
import csv
from scipy.signal import savgol_filter
from PIL import Image, ImageDraw, ImageFont
import random


# -----------------以下为初始配置---------------
warnings.filterwarnings("ignore")  # 忽略警告信息
np.set_printoptions(threshold=np.inf)  # 数据全部打印出来
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False  # 设置画图标题为汉字
font1 = {'family': 'Times New Roman',
         'weight': 'normal',
         'size': 15,
         }  # 英文字体格式
font2 = {'family': 'SimSun',
         'weight': 'normal',
         'size': 15,
         }  # 中文字体格式
ft = ImageFont.truetype('Times New Roman.ttf', 40, encoding="unic")
ft2 = ImageFont.truetype('Times New Roman.ttf', 20, encoding="unic")
colorMap = ['#FF0000', '#00FFFF', '#ADD8E6', '#800080', '#FFFF00', '#00FF00', '#FF00FF', '#FFA500', '#A52A2A', '#800000', '#008000',
    '#808000', '#736AFF', 'pink', '#82CAFF', '#6AFB92', '#D4A017', '#FFCBA4', '#7FFFD4','#FFFFC2']
np.random.seed(2022)
R = np.random.randint(50, 250, 500)  # 生成500个随机数
G = np.random.randint(50, 250, 500)  # 生成500个随机数
B = np.random.randint(50, 250, 500)  # 生成500个随机数

# ----------------------------------------------


def Determine_existence(position, ALL_object, time):  # 寻找position在time帧中的索引号
    """
    :param position: 前一帧的[x,y]
    :param ALL_object: 所有帧所有车辆的[x,y]
    :param time: 第i帧
    :return: 布尔值
    """
    for l in range(len(ALL_object)):
        if ALL_object[l][time] != [] and ALL_object[l][time][0] == position[0] and ALL_object[l][time][1] == position[1]:
            return True, l
    return False, np.inf


def local_assignment(dis_matrix):  # 匈牙利匹配结果
    row_ind = []  # 保存匹配结果
    col_ind = []
    for i in range(min(dis_matrix.shape[0], dis_matrix.shape[1])):  # 循环矩阵的最小维度，假设矩阵为3*5，那么只循环3次，因为最多只能匹配3个目标
        row, col = np.where(dis_matrix == np.min(dis_matrix))  # 求矩阵最小值所在的行号和列号,输出的是一个列表，只有一个值
        if dis_matrix[row[0]][col[0]] > 2:  # 如果矩阵的最小欧式距离都大于2了，说明已经没有可以匹配的对象了，直接return
            row_ind = np.array(row_ind)
            col_ind = np.array(col_ind)
            return row_ind, col_ind
        else:
            row_ind.append(row[0])  # 将最小值的行号和列号作为匹配结果
            col_ind.append(col[0])
            dis_matrix[row, :] = np.inf
            dis_matrix[:, col] = np.inf


def Inspection_assignment(dis_matrix, row_ind, col_ind):  # 检测匈牙利匹配算法是否合适，如果存在匹配结果欧式距离大于2的情况则认为匹配不通过
    flag = 1  # 默认检测通过
    for d in range(len(row_ind)):
        if dis_matrix[row_ind[d], col_ind[d]] > 2:  # 认为欧式距离大于2，匹配失败
            flag = 0  # 检测未通过
            # break
    return flag


def find_col_ind_err(col_ind, new_xy_1):  # 返回当前帧中未被匹配上的物体
    col_ind_err = []  # 当前帧中所有没有被匹配上的物体
    obj_number_1 = 1 if new_xy_1.ndim == 1 else len(new_xy_1)  # 获取当前帧检测到的目标数量
    for i in range(obj_number_1):  # 循环每一个当前物体，判断当前物体是否已经被匹配
        if i not in col_ind:
            col_ind_err.append(i)
    return col_ind_err


def find_start_row_ind_err(row_ind, new_xy_0):  # 返回前一帧中未被匹配上的物体
    row_ind_err = []  # 前一帧中所有没有被匹配上的物体
    obj_number_0 = 1 if new_xy_0.ndim == 1 else len(new_xy_0)  # 获取前一帧检测到的目标数量
    for i in range(obj_number_0):  # 循环每一个前一帧中的物体，判断当前物体是否已经被匹配
        if i not in row_ind:
            row_ind_err.append(i)
    return row_ind_err


def better_KM(new_xy_0, new_xy_1, ALL_object, current, ALL_object_xyzwlh, new_xyzwlh_0, new_xyzwlh_1):  # 当前帧与前一帧中均存在2个以上的目标，使用此函数进行目标匹配与轨迹提取
    """
    :param new_xy_0: 前一帧检测到的目标
    :param new_xy_1: 当前帧检测到的目标
    :param ALL_object: 目标——帧 位置矩阵
    :param current: 当前帧索引号，需要对当前帧进行操作
    :return:ALL_object: 所有目标的轨迹
    """
    dis_matrix = distance.cdist(new_xy_0, new_xy_1, 'euclidean')  # 得到欧式距离矩阵
    row_ind, col_ind = linear_sum_assignment(dis_matrix)  # 首先进行匈牙利全局匹配，得到匹配结果
    # --------得到匹配结果后进行检测--------
    flag = Inspection_assignment(dis_matrix, row_ind, col_ind)  # 返回布尔值，检测通过返回1，未通过返回0
    if not flag:  # 如果检测未通过，进行局部最优匹配
        row_ind, col_ind = local_assignment(dis_matrix)  # 返回局部最优匹配结果
    # col_ind为当前帧中已经匹配上的结果，这些匹配上的结果已经是旧物体。
    # 剩下未匹配上的结果作为新物体
    col_ind_err = find_col_ind_err(col_ind, new_xy_1)  # 当前帧中新物体的索引号

    if current == 1:  # 如果是第1帧，那么需要将第0帧的坐标一并添加到
        row_ind_err = find_start_row_ind_err(row_ind, new_xy_0)  # 初始帧中新物体的索引号
        if len(row_ind_err) > 0:
            for j in range(len(row_ind_err)):  # 循环添加每一个物体的坐标
                ALL_object.append([[] for a in range(frame + 1)])  # 新增一个物体
                ALL_object[-1][current-1] = new_xy_0[row_ind_err[j]]  # 当前帧中第j个新增物体的坐标

                ALL_object_xyzwlh.append([[] for a in range(frame + 1)])  # 新增一个物体
                ALL_object_xyzwlh[-1][current-1] = new_xyzwlh_0[row_ind_err[j]]  # 当前帧中第j个新增物体的坐标

        for j in range(len(row_ind)):  # 循环每个物体的坐标
            '''
             new_xy_0[row_ind[j]]前一帧第j个物体的xy坐标
             new_xy_1[col_ind[j]]当前帧第j个物体的xy坐标
            '''
            ALL_object.append([[] for a in range(frame + 1)])  # 新增一个物体
            ALL_object[-1][1] = new_xy_1[col_ind[j]]  # 当前帧(第1帧)中第j个新增物体的坐标
            ALL_object[-1][0] = new_xy_0[row_ind[j]]  # 补充第0帧中第j个新增物体的坐标

            ALL_object_xyzwlh.append([[] for a in range(frame + 1)])  # 新增一个物体
            ALL_object_xyzwlh[-1][1] = new_xyzwlh_1[col_ind[j]]  # 当前帧(第1帧)中第j个新增物体的坐标
            ALL_object_xyzwlh[-1][0] = new_xyzwlh_0[row_ind[j]]  # 补充第0帧中第j个新增物体的坐标

    else:  # 如果不是初始帧
        for j in range(len(row_ind)):
            # 判断第j个物体的前一帧坐标是否已经存在？ 如果存在，则说明是老物体,返回ALL_object的第l行；否则为新物体，返回np.inf
            boole, l = Determine_existence(new_xy_0[row_ind[j]], ALL_object, current-1)

            if boole:  # 如果是老物体
                ALL_object[l][current] = new_xy_1[col_ind[j]]  # 将第j个物体的坐标添加
                ALL_object_xyzwlh[l][current] = new_xyzwlh_1[col_ind[j]]  # 将第j个物体的坐标添加
            else:  # 如果是新物体，都已经匹配上前一帧了，怎么可能时新物体呢？？
                print(current, "帧发现一个问题，请检查", )

    # 前面添加的是已经匹配上的物体，现在将未匹配上的物体作为新物体添加上
    if len(col_ind_err) > 0:
        for j in range(len(col_ind_err)):  # 循环添加每一个物体的坐标
            ALL_object.append([[] for a in range(frame + 1)])  # 新增一个物体
            ALL_object[-1][current] = new_xy_1[col_ind_err[j]]  # 当前帧中第j个新增物体的坐标

            ALL_object_xyzwlh.append([[] for a in range(frame + 1)])  # 新增一个物体
            ALL_object_xyzwlh[-1][current] = new_xyzwlh_1[col_ind_err[j]]  # 当前帧中第j个新增物体的坐标

    return ALL_object, ALL_object_xyzwlh


def Euclidean_distance(position_1, position_2):  # 计算两个目标的欧式距离
    x1, y1 = position_1[0], position_1[1]
    x2, y2 = position_2[0], position_2[1]
    return math.\
        sqrt(((x2 - x1) ** 2 + (y2 - y1)) ** 2)


def assignment_1(new_xy_0, new_xy_1):  # 当前帧为1个目标 或 前一帧为1个目标，进行匹配，返回匹配结果
    obj_number_0 = 1 if new_xy_0.ndim == 1 else len(new_xy_0)  # 获取前一帧(i)检测到的目标数量
    obj_number_1 = 1 if new_xy_1.ndim == 1 else len(new_xy_1)  # 获取当前帧(i+1)检测到的目标数量
    row_ind = []
    col_ind = []
    col_ind_err = []  # 当前帧匹配失败的物体认为是新增物体
    distance_1 = []
    if obj_number_0 == 1 and obj_number_1 == 1:  # 如果前后两帧均为1个目标
        distance_1.append(Euclidean_distance(new_xy_0, new_xy_1))
        if min(distance_1) <= 2:  # 如果最小欧氏距离小于2，说明匹配成功。
            row_ind.append(0)  # 索引号只可能为1
            col_ind.append(0)  # 添加匹配成功的索引号
        else:  # 匹配不成功
            col_ind_err.append(0)

    elif obj_number_0 == 1:  # 如果前1帧为1个目标
        for i in range(obj_number_1):
            distance_1.append(Euclidean_distance(new_xy_0, new_xy_1[i]))

        if min(distance_1) <= 2:  # 如果最小欧氏距离小于2，说明匹配成功
            ind = distance_1.index(min(distance_1))
            row_ind.append(0)
            col_ind.append(ind)  # 添加匹配成功的索引号
            for k in range(obj_number_1):
                if k != ind:
                    col_ind_err.append(k)  # 添加匹配失败的索引号
        else:  # 如果匹配失败
            for k in range(obj_number_1):
                 col_ind_err.append(k)  # 匹配失败需要添加当前帧所有物体的索引号，作为新增物体

    else:  # 如果当前帧为1个目标
        for i in range(obj_number_0):
            distance_1.append(Euclidean_distance(new_xy_0[i], new_xy_1))

        if min(distance_1) <= 2:  # 如果最小欧氏距离小于2，说明匹配成功
            ind = distance_1.index(min(distance_1))
            row_ind.append(ind)
            col_ind.append(0)
        else:
            col_ind_err.append(0)  # 当前帧为1个目标且匹配失败，那么当前这个物体为新增物体

    return row_ind, col_ind, col_ind_err


def better_method_1(new_xy_0, new_xy_1, ALL_object, current, ALL_object_xyzwlh, new_xyzwlh_1):  # 当前帧或者上一帧只检测到一个物体用这个函数
    """
    :param new_xy_0: 前一帧检测到的目标
    :param new_xy_1: 当前帧检测到的目标
    :param ALL_object: 目标——帧 位置矩阵
    :param current: 当前帧索引号，需要对当前帧进行操作
    :return:ALL_object: 所有目标的轨迹
    """
    row_ind, col_ind, col_ind_err = assignment_1(new_xy_0, new_xy_1)
    if len(row_ind) != 0:  # 如果匹配成功
        if new_xy_0.ndim == 1:  # 如果前面一帧只有1个维度，那么直接输入
            boole, l = Determine_existence(new_xy_0, ALL_object, current - 1)  # 这是匹配上的索引号，一定是旧物体
        else:  # 如果前面一帧有多个维度
            boole, l = Determine_existence(new_xy_0[row_ind[0]], ALL_object, current - 1)  # 这是匹配上的索引号，一定是旧物体

        if new_xy_1.ndim == 1:  # 如果当前帧只有1个维度
            ALL_object[l][current] = new_xy_1
            ALL_object_xyzwlh[l][current] = new_xyzwlh_1
        else:
            ALL_object[l][current] = new_xy_1[col_ind[0]]  # 将第j个物体的坐标添加当前帧
            ALL_object_xyzwlh[l][current] = new_xyzwlh_1[col_ind[0]]

    # 如果当前帧有匹配失败的物体，则作为新增物体
    if len(col_ind_err) > 0:
        for j in range(len(col_ind_err)):  # 循环添加每一个物体的坐标
            ALL_object.append([[] for a in range(frame + 1)])  # 新增一个物体
            ALL_object[-1][current] = new_xy_1[col_ind_err[j]]  # 当前帧中第j个新增物体的坐标

            ALL_object_xyzwlh.append([[] for a in range(frame + 1)])  # 新增一个物体
            ALL_object_xyzwlh[-1][current] = new_xyzwlh_1[col_ind_err[j]]  # 当前帧中第j个新增物体的坐标

    return ALL_object,  ALL_object_xyzwlh


def Srotation_angle_get_coor_coordinates(point, center, angle):  # 坐标系转换
    # 逆时针旋转
    src_x, src_y = point
    center_x, center_y = center
    radian = math.radians(angle)
    dest_x = round(((src_x - center_x) * math.cos(radian) - (src_y - center_y) * math.sin(radian) + center_x), 3)
    dest_y = round(((src_x - center_x) * math.sin(radian) + (src_y - center_y) * math.cos(radian) + center_y), 3)
    return [dest_x, dest_y]


def plot_Animated(ALL_object):  # 记录每一帧每个目标的位置，进行可视化
    # 定义圆圈
    circle_r = 500  # 画半径为200的圆
    for Frame in range(len(ALL_object[1])):  # 循环每一帧
        circle_Y = []
        circle_X = []
        obj_id = []
        for i in range(len(ALL_object)):  # 循环每一个目标
            obj_id.append(i)
            if len(ALL_object[i][Frame]) != 0:  # 判断当前目标是否在当前帧中有坐标点
                circle_x, circle_y = ALL_object[i][Frame][0], ALL_object[i][Frame][1]
                circle_Y.append(circle_y)
                circle_X.append(circle_x)
            else:
                circle_Y.append(-100)  # 保证维度相同，所需必须要添加坐标
                circle_X.append(-100)

        # 下面是画图部分
        plt.scatter(circle_X, circle_Y, c=obj_id, s=circle_r, cmap=plt.cm.Set1)  # 用不同颜色画每一个点

        for j in range(len(obj_id)):  # 给每一个目标添加编号
            plt.text(circle_X[j],
                     circle_Y[j],
                     str(obj_id[j]),
                     fontsize=15,
                     verticalalignment="center",
                     horizontalalignment="center"
                     )
        # 画矩形框
        left, right, bottom, top = [-25, 25, 4, 40]
        rectangle_x = [left, left, right, right, left]
        rectangle_y = [top, bottom, bottom, top, top]
        plt.plot(rectangle_x, rectangle_y, 'k--')
        plt.xlabel("X")
        plt.ylabel("Y")
        plt.xlim(-40, 40)
        plt.ylim(0, 50)
        plt.savefig(os.getcwd() + '\\output\\image_test' + str(Frame) + '.png')
        # plt.show()  # 或者用下面三行代码，快速保存
        plt.ion()
        plt.pause(0.1)
        plt.close()


def plot_image(ALL_object):
    for i in range(len(ALL_object)):
        X = []
        Y = []
        for j in range(len(ALL_object[i])):
            if len(ALL_object[i][j]) != 0:
                X.append(ALL_object[i][j][0])
                Y.append(ALL_object[i][j][1])
        plt.plot(X, Y, c=colorMap[i], linewidth=2)
    plt.yticks(fontproperties='Times New Roman', size=15)  # 设置大小及加粗
    plt.xticks(fontproperties='Times New Roman', size=15)
    plt.xlim(-25,25)  # 去x坐标刻度
    plt.ylim(-5,45)  # 去y坐标刻度
    plt.xlabel("X", font1)
    plt.ylabel("Y", font1)
    plt.scatter(0,0,marker='p', s=50, color='r', label="LiDAR")
    plt.legend(loc='best',facecolor='#FFF0F5', prop=font1)
    plt.savefig('test.svg', bbox_inches='tight')
    plt.show()


def plot_image_Animated(ALL_object):
    # ---------------------画圆---------------------
    circle_r = 500  # 画半径为200的圆
    for Frame in range(len(ALL_object[1])):  # 循环每一帧
        circle_Y = []
        circle_X = []
        obj_id = []
        for i in range(len(ALL_object)):  # 循环每一个目标
            obj_id.append(i)
            if len(ALL_object[i][Frame]) != 0:  # 判断当前目标是否在当前帧中有坐标点
                circle_x, circle_y = ALL_object[i][Frame][0], ALL_object[i][Frame][1]
                circle_Y.append(circle_y)
                circle_X.append(circle_x)
            else:
                circle_Y.append(-100)  # 保证维度相同，所需必须要添加坐标
                circle_X.append(-100)

        # 下面是画图部分

        plt.scatter(circle_X, circle_Y, c=obj_id, s=circle_r, cmap=plt.cm.Set1)  # 用不同颜色画每一个点
        for j in range(len(obj_id)):  # 给每一个目标添加编号
            plt.text(circle_X[j],
                     circle_Y[j],
                     str(obj_id[j]),
                     fontsize=15,
                     verticalalignment="center",
                     horizontalalignment="center"
                     )

        # ----------------------画轨迹--------------------------
        for i in range(len(ALL_object)):     # 循环每一个目标
            X = []
            Y = []
            for j in range(Frame):
                if len(ALL_object[i][j]) != 0:
                    X.append(ALL_object[i][j][0])
                    Y.append(ALL_object[i][j][1])
                    plt.plot(X, Y, c=colorMap[i], linewidth=2)

        plt.yticks(fontproperties='Times New Roman', size=15)  # 设置大小及加粗
        plt.xticks(fontproperties='Times New Roman', size=15)
        plt.xlim(-25,25)  # 去x坐标刻度
        plt.ylim(-5,45)  # 去y坐标刻度
        plt.xlabel("X", font1)
        plt.ylabel("Y", font1)
        plt.scatter(0,0,marker='p', s=50, color='r', label="LiDAR")
        plt.legend(loc='best',facecolor='#FFF0F5',prop=font1)
        plt.savefig(os.getcwd() + '\\output\\image_test\\' + str(Frame) + '.png')
        # plt.show()  # 或者用下面三行代码，快速保存
        plt.ion()
        plt.pause(0.1)
        plt.close()


def Coordinate_conversion(x, y, im):
    # LiDAR坐标系坐标转为im坐标
    x_new = x * 12.17 + im.size[0]/2
    y_new = im.size[1] - y * 12.19
    return x_new, y_new


def plot_in_image(ALL_object):
    circle_r = 20  # 画半径为20像素的圆
    for Frame in range(len(ALL_object[1])):  # 循环每一帧
        imagePath = r'E:\\Study\\jason\\CIITR-LiDAR-main\\New Dev\\Track_SGH\\object_detection\\test' + str(Frame) +'.png'  # 读取每一帧作为背景
        im = Image.open(imagePath)  # 加载背景
        draw = ImageDraw.Draw(im)
        circle_XY = []  # 存储圆坐标
        text_id = []   # 存储文本坐标，即id编号坐标
        text_xyzwlh = []   # 存储目标中心点xyz及lwh坐标

        obj_id = []  # 存储圆编号
        obj_xyzwlh = []  # 存储xyzwlh文本值
        for i in range(len(ALL_object)):  # 循环每一个目标
            obj_id.append(i)
            temp = []
            if len(ALL_object[i][Frame]) != 0:  # 判断当前目标是否在当前帧中有坐标点
                circle_x, circle_y = Coordinate_conversion(ALL_object[i][Frame][0], ALL_object[i][Frame][1], im)
                x1 = circle_x - circle_r
                y1 = circle_y - circle_r
                x2 = circle_x + circle_r
                y2 = circle_y + circle_r
                temp.append((x1, y1))
                temp.append((x2, y2))
                circle_XY.append(temp)
                text_id.append((x1+10, y1))
                text_xyzwlh.append((x1-10, y1-130))
                obj_xyzwlh.append(ALL_object[i][Frame])
            else:
                temp.append((-100, -100))
                temp.append((-200, -200))
                circle_XY.append(temp)
                text_id.append((-100, -100))
                text_xyzwlh.append((-100, -100))
                obj_xyzwlh.append(np.array([1, 1, 1, 1, 1, 1]))  # 保持维度相同

        # ----------------------画轨迹--------------------------
        for i in range(len(ALL_object)):     # 循环每一个目标
            xy = []
            for j in range(Frame):
                if len(ALL_object[i][j]) != 0:
                    x_new, y_new = Coordinate_conversion(ALL_object[i][j][0], ALL_object[i][j][1], im)
                    xy.append(x_new)
                    xy.append(y_new)
            draw.line(xy, fill=(R[i], G[i], B[i]), width=5)  # 线的起点和终点，线宽

        # ----------------------画文字--------------------------
        for id in range(len(circle_XY)):
            draw.chord(circle_XY[id], 0, 360, outline=(R[obj_id[id]], G[obj_id[id]], B[obj_id[id]]), width=2)
            draw.text(text_id[id], str(obj_id[id]), fill=(R[obj_id[id]], G[obj_id[id]], B[obj_id[id]]), font=ft)

            text_temp = " x:" + str(obj_xyzwlh[id][0]) + "\n y:" + str(obj_xyzwlh[id][1]) + "\n z:" + str(obj_xyzwlh[id][2]) + \
                        "\n w:" + str(obj_xyzwlh[id][3]) + "\n l:" + str(obj_xyzwlh[id][4]) + "\n h:" + str(obj_xyzwlh[id][5])
            draw.text(text_xyzwlh[id], text_temp, fill=(R[obj_id[id]], G[obj_id[id]], B[obj_id[id]]), font=ft2)

        # im.show()
        im.save(r'E:\\Study\\jason\\CIITR-LiDAR-main\\New Dev\\Track_SGH\\output/image_test/' + str(Frame) + '.png')


def plot_speed_smooth(speed):
    smooth_speed = np.zeros((speed.shape[0], speed.shape[1]))
    for i in range(speed.shape[0]):
        for j in range(len(speed[i])):
            if np.isnan(speed[i][j]) == 0 and np.isnan(speed[i][j+1]) == 1:  # 当前j为有值，下一个j无值，右边界
                temp = -(speed[i][j-2] - speed[i][j-1])/speed[i][j] + speed[i][j-1]
                smooth_speed[i][j] = temp
            elif np.isnan(speed[i][j]) == 0 and np.isnan(speed[i][j-1]) == 1:  # 当前j为有值，上一个j无值，左边界
                temp = -(speed[i][j+1] - speed[i][j+2])/speed[i][j] + speed[i][j+1]
                smooth_speed[i][j] = temp
            elif 2 <= j < len(speed[i])-2 and not np.isnan(speed[i][j-2]) and not np.isnan(speed[i][j+2]):
                temp = (speed[i][j-2] + speed[i][j-1] + speed[i][j] + speed[i][j+2] + speed[i][j+1]) / 5
                smooth_speed[i][j] = temp
            else:
                smooth_speed[i][j] = speed[i][j]
        plt.plot(smooth_speed[i],c=colorMap[i], linewidth=2)
    plt.xlim(0, 120)  # 去x坐标刻度
    plt.ylim(0, 20)  # 去y坐标刻度
    plt.yticks(fontproperties='Times New Roman', size=15)  # 设置大小及加粗
    plt.xticks(fontproperties='Times New Roman', size=15)
    plt.xlabel("Frame", font1)
    plt.ylabel("velocity(m/s)", font1)
    plt.show()


def plot_speed(speed):
    for i in range(speed.shape[0]):
        plt.plot(speed[i],c=colorMap[i], linewidth=2)
    plt.xlim(0, 120)  # 去x坐标刻度
    plt.ylim(0, 20)  # 去y坐标刻度
    plt.yticks(fontproperties='Times New Roman', size=15)  # 设置大小及加粗
    plt.xticks(fontproperties='Times New Roman', size=15)
    plt.xlabel("Frame", font1)
    plt.ylabel("Velocity(m/s)", font1)
    plt.show()


def write_data(ALL_object):  # 写入轨迹提取后所有目标的轨迹坐标
    filename = os.getcwd() + '\\output\\track.xlsx'
    workbook = openpyxl.load_workbook(filename)
    Sheet1 = workbook["Sheet1"]
    for n in range(len(ALL_object)):
        for f in range(len(ALL_object[n])):
            Sheet1.cell(n+1, f+1).value = str(ALL_object[n][f])
    workbook.save(filename)


def write_merger(merger):  # 写入合并目标组合
    filename = os.getcwd() + '\\output\\track.xlsx'
    workbook = openpyxl.load_workbook(filename)
    Sheet2 = workbook["Sheet2"]
    for n in range(len(merger)):
        for f in range(len(merger[n])):
            Sheet2.cell(n + 1, f + 1).value = merger[n][f]
    workbook.save(filename)


def write_new_ALL_object(new_ALL_object):  # 写入轨迹合并、补全后所有目标的轨迹坐标
    filename = os.getcwd() + '\\output\\track.xlsx'
    workbook = openpyxl.load_workbook(filename)
    Sheet3 = workbook["Sheet3"]
    for n in range(len(new_ALL_object)):
        for f in range(len(new_ALL_object[n])):
            Sheet3.cell(n + 1, f + 1).value = str(new_ALL_object[n][f])
    workbook.save(filename)


def determine_breakpoint(ALL_object):
    breakpoint_idx = []   # 0表示轨迹可能完整，1表示轨迹一定不完整
    Frame_number = 20   # 认为小于20帧，这个轨迹不完整
    for n in range(len(ALL_object)):
        count = 0
        for f in range(len(ALL_object[n])):
            if len(ALL_object[n][f]) != 0:
                count += 1
        if count < Frame_number:
            breakpoint_idx.append(1)
        else:
            breakpoint_idx.append(0)
    return breakpoint_idx


def start_end(ALL_object, n, flag):  # 寻找第n辆车的起始点/结束点,   flag = 1，求起始点。 flag = -1 ，求结束点
    if flag == 1:   # 求起始点
        start_xy = []
        for i in range(len(ALL_object[n])):
            if i == 0:
                if len(ALL_object[n][i]) != 0:   # 这个目标从激光雷达扫描开始就已经存在了
                    start_xy.append(ALL_object[n][i][0])
                    start_xy.append(ALL_object[n][i][1])
                    return start_xy, i
            elif len(ALL_object[n][i-1]) == 0 and len(ALL_object[n][i]) != 0:  # 前一帧没有轨迹，这一帧有轨迹。说明这一帧是起始点
                start_xy.append(ALL_object[n][i][0])
                start_xy.append(ALL_object[n][i][1])
                return start_xy, i
        print("未找到起始点")

    elif flag == -1:   # 求结束点
        end_xy = []
        for i in range(len(ALL_object[n])):
            if i != len(ALL_object[n])-1:
                if len(ALL_object[n][i]) != 0 and len(ALL_object[n][i+1]) == 0:  # 当前帧有轨迹，后一帧无轨迹。说明当期帧是结束点
                    end_xy.append(ALL_object[n][i][0])
                    end_xy.append(ALL_object[n][i][1])
                    return end_xy, i
            elif len(ALL_object[n][i]) != 0:   # 在激光雷达断电时，这个物体还存在
                end_xy.append(ALL_object[n][i][0])
                end_xy.append(ALL_object[n][i][1])
                return end_xy, i
        print("未找到结束点")

    else:
        print("输入的flag应该为1或-1")


def find_speed(start_xy, end_xy, start_frame, end_frame):  # 计算在固定帧间隔下目标的速度
    time = (end_frame - start_frame) / 10    # 假设激光雷达转速为每秒10帧
    distance = math.pow(math.pow(start_xy[0] - end_xy[0], 2) + math.pow(start_xy[1] - end_xy[1], 2), 0.5)
    speed = round(distance/time, 2)   # 单位为   m/s
    return speed


def find_attr(ALL_object):  # 记录当前每个目标的属性
    track_attribute = []  # 1.方向(-1,0,1)   2.速度(单帧情况下速度为None)   3.起始坐标   4.起始帧    5.结束坐标    6.结束帧
    for n in range(len(ALL_object)):  # 0表示轨迹可能完整，1表示轨迹一定不完整
        start_xy, start_frame = start_end(ALL_object, n, 1)  # 求第n个目标的起点坐标
        end_xy, end_frame = start_end(ALL_object, n, -1)  # 求第n个目标的终点坐标
        if end_frame - start_frame >= 1:  # 两帧及以上可以计算方向、速度
            if start_xy[0] < end_xy[0]:  # 说明是从左向右运动
                direction = 1
            else:
                direction = -1  # 说明是从右向左运动
            speed = find_speed(start_xy, end_xy, start_frame, end_frame)
            attribute = [direction, speed, start_xy, start_frame, end_xy, end_frame]
            track_attribute.append(attribute)
        else:  # 否则就是单帧情况
            attribute = [None, None, start_xy, start_frame, end_xy, end_frame]
            track_attribute.append(attribute)
    return track_attribute


def position_x_Similarity(xn, xt, frame_n, frame_t, speed_n, direction_n):  # 判断预测值与实际值的轨迹相似性，并返回误差值，如果相似，则说明是一个物体

    """
    判断xn和xt的相似性，返回布尔值。如果返回值为真，可以进行匹配
    :param xn: 第n个目标的x坐标
    :param xt: 待匹配目标的x坐标
    :param frame_n: 第n个目标的帧
    :param frame_t: 待匹配目标的帧
    :param speed_n: 第n个目标的速度
    :param direction_n: 第n个目标的行驶方向
    :return: 布尔值, 轨迹预测误差  越接近0，说明误差越小，＜2时，检测通过
    """
    Threshold = 2
    if frame_t < frame_n:  # 说明是寻找第n个目标的上游
        if direction_n == 1:  # 说明是从左向右匹配，即计算xn左边第（frame_n - frame_t）帧的坐标，看是否与xt相似
            new_xn = xn - speed_n * (frame_n - frame_t)/10
            if abs(new_xn - xt) < Threshold:
                return True, round(abs(new_xn - xt), 3)
            else:
                return False, round(abs(new_xn - xt), 3)
        elif direction_n == -1:   # 说明是从右向左匹配，即计算xn右边第（frame_n - frame_t）帧的坐标
            new_xn = xn + speed_n * (frame_n - frame_t)/10
            if abs(new_xn - xt) < Threshold:
                return True, round(abs(new_xn - xt), 3)
            else:
                return False, round(abs(new_xn - xt), 3)
        else:
            print("xn必须要有方向，要么为1，要么为-1")
    elif frame_t > frame_n:  # 说明是寻找第n个目标的下游
        if direction_n == 1:  # 说明是从左向右匹配，即计算xn右边第（frame_t - frame_n）帧的坐标
            new_xn = xn + speed_n * (frame_t - frame_n) / 10
            if abs(new_xn - xt) < Threshold:
                return True, round(abs(new_xn - xt), 3)
            else:
                return False, round(abs(new_xn - xt), 3)
        elif direction_n == -1:  # 说明是从右向左匹配，即计算xn左边第（frame_t - frame_n）帧的坐标
            new_xn = xn - speed_n * (frame_t - frame_n) / 10
            if abs(new_xn - xt) < Threshold:
                return True, round(abs(new_xn - xt), 3)
            else:
                return False, round(abs(new_xn - xt), 3)
    else:
        print("frame_t 不能等于 frame_n")


def exist_n_t(merger, n, t):   # 若只有n存在，则返回n存在的索引号和1；  若只有t存在，则返回t存在的索引号和-1   若n和t都存在，则返回索引号和0；   若n和t都不存在，则返回None和None
    flag = None  # 假设n和t都不存在
    for i in range(len(merger)):
        if n in merger[i] and t in merger[i]:  # 如果t和n都在这个merger[i]这个列表内
            flag = 0
            return i, flag
        elif n in merger[i] and t not in merger[i]:
            flag = 1
            return i, flag
        elif n not in merger[i] and t in merger[i]:
            flag = -1
            return i, flag
    return None, flag   # 当merger循环完后，依旧没有找到前三种情况，说明n和t都不存在


def filled_xy(xy_incomplete):  # 对目标缺失轨迹进行补全
    start_xy = []    # 记录插值的起点坐标
    end_xy = []      # 记录插值的终点坐标
    start_f = []     # 记录插值的起点帧
    end_f = []       # 记录插值的终点帧
    for i in range(len(xy_incomplete)):
        if i < len(xy_incomplete)-1:
            if len(xy_incomplete[i]) != 0 and len(xy_incomplete[i+1]) == 0:   # 第i帧有值，i+1帧没有值，记录i
                start_xy.append(xy_incomplete[i])
                start_f.append(i)
            if len(xy_incomplete[i]) == 0 and len(xy_incomplete[i+1]) != 0:
                end_xy.append(xy_incomplete[i+1])
                end_f.append(i+1)
    # print("***************")
    # print("start_xy:", start_xy)
    # print("end_xy:", end_xy)
    # print("start_f:", start_f)
    # print("end_f:", end_f)
    # 插值补全
    for ind in range(len(start_xy)):
        for f in range(start_f[ind]+1, end_f[ind], 1):
            x_incomplete = (end_xy[ind][0] - start_xy[ind][0]) / (end_f[ind] - start_f[ind]) * (f - start_f[ind]) + start_xy[ind][0]   # (Xend - Xstart) / (Fend - Fstart) + Xstart
            y_incomplete = (end_xy[ind][1] - start_xy[ind][1]) / (end_f[ind] - start_f[ind]) * (f - start_f[ind]) + start_xy[ind][1]   # (Yend - Ystart) / (Fend - Fstart) + Ystart
            xy_incomplete[f] = [round(x_incomplete, 3), round(y_incomplete, 3)]
    return xy_incomplete


def filled_xyzwlh(xyzwlh_incomplete):  # 对目标缺失轨迹进行补全
    start_xy = []    # 记录插值的起点坐标
    end_xy = []      # 记录插值的终点坐标
    start_f = []     # 记录插值的起点帧
    end_f = []       # 记录插值的终点帧
    for i in range(len(xyzwlh_incomplete)):
        if i < len(xyzwlh_incomplete)-1:
            if len(xyzwlh_incomplete[i]) != 0 and len(xyzwlh_incomplete[i+1]) == 0:   # 第i帧有值，i+1帧没有值，记录i
                start_xy.append(xyzwlh_incomplete[i])
                start_f.append(i)
            if len(xyzwlh_incomplete[i]) == 0 and len(xyzwlh_incomplete[i+1]) != 0:
                end_xy.append(xyzwlh_incomplete[i+1])
                end_f.append(i+1)
    # print("***************")
    # print("start_xy:", start_xy)
    # print("end_xy:", end_xy)
    # print("start_f:", start_f)
    # print("end_f:", end_f)
    # 插值补全
    for ind in range(len(start_xy)):
        for f in range(start_f[ind]+1, end_f[ind], 1):
            x_incomplete = (end_xy[ind][0] - start_xy[ind][0]) / (end_f[ind] - start_f[ind]) * (f - start_f[ind]) + start_xy[ind][0]   # (Xend - Xstart) / (Fend - Fstart) + Xstart
            y_incomplete = (end_xy[ind][1] - start_xy[ind][1]) / (end_f[ind] - start_f[ind]) * (f - start_f[ind]) + start_xy[ind][1]   # (Yend - Ystart) / (Fend - Fstart) + Ystart
            z_incomplete = (end_xy[ind][2] - start_xy[ind][2]) / (end_f[ind] - start_f[ind] + 0.000000001) * (f - start_f[ind]) + start_xy[ind][2]   # 防止分母为0
            w_incomplete = (end_xy[ind][3] - start_xy[ind][3]) / (end_f[ind] - start_f[ind] + 0.000000001) * (f - start_f[ind]) + start_xy[ind][3]
            l_incomplete = (end_xy[ind][4] - start_xy[ind][4]) / (end_f[ind] - start_f[ind] + 0.000000001) * (f - start_f[ind]) + start_xy[ind][4]
            h_incomplete = (end_xy[ind][5] - start_xy[ind][5]) / (end_f[ind] - start_f[ind] + 0.000000001) * (f - start_f[ind]) + start_xy[ind][5]
            xyzwlh_incomplete[f] = [round(x_incomplete, 3), round(y_incomplete, 3), round(z_incomplete, 3), round(w_incomplete, 3), round(l_incomplete, 3), round(h_incomplete, 3)]
    return xyzwlh_incomplete


def Calculation_speed(ALL_object):
    row = len(ALL_object)
    col = len(ALL_object[0])
    speed = np.zeros((row, col))
    time = 0.1  # 每帧时间间隔为0.1秒
    for i in range(row):
        for j in range(col):
            if len(ALL_object[i][j-1]) != 0 and 1 <= j <= col-1 and len(ALL_object[i][j+1]) != 0:
                x1 = ALL_object[i][j-1][0]
                y1 = ALL_object[i][j-1][1]
                x2 = ALL_object[i][j+1][0]
                y2 = ALL_object[i][j+1][1]
                v_temp = math.pow(math.pow(y2 - y1, 2) + math.pow(x2 - x1, 2), 0.5) / (time*2)
                speed[i][j] = round(v_temp, 2)
    for i in range(speed.shape[0]):
        for j in range(speed.shape[1]):
            if speed[i][j] == 0:
                speed[i][j] = None
    return speed


if __name__ == '__main__':
    # ------------------------------------初始设置------------------------------------
    file_path = os.getcwd()
    input_path = file_path + '\\input_xyzwlh\\'

    frame = 146
    ALL_object = []  # 存储目标——帧 位置矩阵
    ALL_object_xyzwlh = []  # 只存储(目标中心点的xyz和长宽高)，不进行计算
    # ------------------------------------轨迹提取------------------------------------
    for i in range(frame):
        current = i + 1  # 当前帧
        before = i  # 上一帧
        new_xy_0 = np.loadtxt(input_path + str(before) + '.csv', delimiter=",", usecols=(1, 2), unpack=True, encoding='UTF-8')
        new_xy_1 = np.loadtxt(input_path + str(current) + '.csv', delimiter=",", usecols=(1, 2), unpack=True, encoding='UTF-8')

        new_xyzwlh_0 = np.loadtxt(input_path + str(before) + '.csv', delimiter=",", usecols=(1, 2, 3, 4, 5, 6), unpack=True, encoding='UTF-8')
        new_xyzwlh_1 = np.loadtxt(input_path + str(current) + '.csv', delimiter=",", usecols=(1, 2, 3, 4, 5, 6), unpack=True, encoding='UTF-8')

        new_xyzwlh_0 = np.transpose(new_xyzwlh_0)
        new_xyzwlh_1 = np.transpose(new_xyzwlh_1)

        if len(new_xy_1) == 0:
            continue
        else:
            new_xy_1 = np.transpose(new_xy_1)  # 当前帧的XY矩阵
            obj_number_1 = 1 if new_xy_1.ndim == 1 else len(new_xy_1)  # 获取当前帧(i+1)检测到的目标数量

        if len(new_xy_0) != 0:
            new_xy_0 = np.transpose(new_xy_0)  # 前一帧的XY矩阵
            obj_number_0 = 1 if new_xy_0.ndim == 1 else len(new_xy_0)  # 三元运算符，等于下面这个式子
            '''
                   if new_xy_0.ndim == 1:
                       obj_number_0 = 1
                   else:
                       obj_number_0 = len(new_xy_0)  # 获取前一帧(i)检测到的目标数量
            '''
        else:
            obj_number_0 = 0

        if obj_number_1 == 0:  # 如果当前帧没有检测到物体，则不进行操作
            continue

        elif obj_number_0 == 0:  # 如果当前帧检测到物体了，且上一帧未检测到物体，则该物体为新物体
            if obj_number_1 == 1:
                ALL_object.append([[] for a in range(frame + 1)])  # 新增一个物体
                ALL_object[-1][current] = new_xy_1  # 当前帧中第j个新增物体的坐标

                ALL_object_xyzwlh.append([[] for a in range(frame + 1)])
                ALL_object_xyzwlh[-1][current] = new_xyzwlh_1
            else:
                for j in range(obj_number_1):  # 循环添加每一个物体的坐标
                    ALL_object.append([[] for a in range(frame + 1)])  # 新增一个物体
                    ALL_object[-1][current] = new_xy_1[j]  # 当前帧中第j个新增物体的坐标

                    ALL_object_xyzwlh.append([[] for a in range(frame + 1)])
                    ALL_object_xyzwlh[-1][current] = new_xyzwlh_1[j]

        elif obj_number_0 == 1 or obj_number_1 == 1:  # 当前帧或者上一帧只检测到一个物体
            ALL_object, ALL_object_xyzwlh = better_method_1(new_xy_0, new_xy_1, ALL_object, current, ALL_object_xyzwlh, new_xyzwlh_1)

        elif obj_number_0 >= 2 and obj_number_1 >= 2:  # 当前帧和上一帧均检测到两个以上的物体
            ALL_object, ALL_object_xyzwlh = better_KM(new_xy_0, new_xy_1, ALL_object, current, ALL_object_xyzwlh, new_xyzwlh_0, new_xyzwlh_1)

        else:
            print("第", i+1, "帧存在问题.", "前一帧检测到的目标数为：", obj_number_0, "——当前帧检测到的目标数为：", obj_number_1)

    # 激光雷达扫描到的坐标系与实际大地坐标系不一样，需要进行坐标转换
    for n in range(len(ALL_object)):
        for f in range(frame):
            if len(ALL_object[n][f]) != 0:
                new_x, new_y = Srotation_angle_get_coor_coordinates(ALL_object[n][f], [0, 0], 90)  # 需要XY坐标互换
                ALL_object[n][f] = [new_x, new_y]
                ALL_object_xyzwlh[n][f][0] = new_x
                ALL_object_xyzwlh[n][f][1] = new_y

    # ------------------------------------------轨迹跟踪--------------------------------------------------
    track_attribute = find_attr(ALL_object)    # 计算每一个目标的属性  1.方向(-1,None,1)   2.速度(单帧情况下速度为None)   3.起始坐标   4.起始帧    5.结束坐标    6.结束帧
                                                          # index：         0                    1                     2          3         4           5

    merger = []     # 多个目标合并为1个目标的索引号
    for n in range(len(track_attribute)):
        if 1 <= track_attribute[n][-1] - track_attribute[n][-3] < 36:  # <30说明第n个目标的轨迹一定是不完整的, >=1说明至少为两帧，可以计算方向和速度，进行主动匹配

            for t in range(len(track_attribute)):   # 循环每一个目标，进行配对   track_attribute[t]为待匹配目标

                if t != n:   # t和n不相同时才能进行匹配
                    if track_attribute[t][0] == track_attribute[n][0] or track_attribute[t][0] == None:  # 方向必须相同才能配对,要么为1，要么为-1

                        direction = track_attribute[n][0]    # 获取方向
                        if direction == 1:  # 如果是从左向右行驶，那么待匹配目标的终点x坐标应小于第n个目标起点x坐标 且 待匹配目标的终点帧小于第n个目标起点帧  （寻找左侧/上游）
                            # 1.---------------------------------->->->从左向右，匹配上游------------------------------------------
                            if track_attribute[t][4][0] < track_attribute[n][2][0] and track_attribute[t][5] < track_attribute[n][3]:  # 寻找左侧/上游
                                xn = track_attribute[n][2][0]
                                xt = track_attribute[t][4][0]
                                frame_n = track_attribute[n][3]
                                frame_t = track_attribute[t][5]
                                speed_n = track_attribute[n][1]
                                direction_n = track_attribute[n][0]
                                Similarity, err = position_x_Similarity(xn, xt, frame_n, frame_t, speed_n, direction_n)

                                if Similarity:    # 如果为真，说明第n个目标和第t个目标可以合并
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "<2,可以合并")
                                    if len(merger) == 0:   # 说明还没有值
                                        merger.append([n, t])
                                    else:   # 如果Merger已经存在值了，则进行判断
                                        index_m, flag_m = exist_n_t(merger, n, t)     # index 为当前已存在n或者t的索引号，flag为是否存在n或者t
                                        if flag_m == None:
                                            merger.append([n, t])
                                        elif flag_m == 1:   # 说明只有n存在，则将t添加到n存在的索引号
                                            merger[index_m].append(t)
                                        elif flag_m == -1:  # 说明只有t存在，则将n添加到t存在的索引号
                                            merger[index_m].append(n)
                                        else:   # 剩下就是n和t都存在的情况了，这种情况是，判断t的时候将n一起添加了，然后现在到了n的时候，结果寻找到t了
                                            pass
                                else:
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "≥2,不可以合并")
                            # 2.---------------------------------------从左向右，匹配下游--------------------------------------->->->
                            # 待匹配目标的起点x坐标应大于第n个目标终点x坐标 且 待匹配目标的起点帧大于第n个目标终点帧  （寻找右侧/下游）
                            elif track_attribute[t][2][0] > track_attribute[n][4][0] and track_attribute[t][3] > track_attribute[n][5]:  # 寻找右侧/下游
                                xn = track_attribute[n][4][0]
                                xt = track_attribute[t][2][0]
                                frame_n = track_attribute[n][5]
                                frame_t = track_attribute[t][3]
                                speed_n = track_attribute[n][1]
                                direction_n = track_attribute[n][0]
                                Similarity, err = position_x_Similarity(xn, xt, frame_n, frame_t, speed_n, direction_n)
                                if Similarity:    # 如果为真，说明第n个目标和第t个目标可以合并
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "<2,可以合并")
                                    if len(merger) == 0:   # 说明还没有值
                                        merger.append([n, t])
                                    else:   # 如果Merger已经存在值了，则进行判断
                                        index_m, flag_m = exist_n_t(merger, n, t)     # index 为当前已存在n或者t的索引号，flag为是否存在n或者t
                                        if flag_m == None:
                                            merger.append([n, t])
                                        elif flag_m == 1:   # 说明只有n存在，则将t添加到n存在的索引号
                                            merger[index_m].append(t)
                                        elif flag_m == -1:  # 说明只有t存在，则将n添加到t存在的索引号
                                            merger[index_m].append(n)
                                        else:   # 剩下就是n和t都存在的情况了，这种情况是，判断t的时候将n一起添加了，然后现在到了n的时候，结果寻找到t了
                                            pass
                                else:
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "≥2,不可以合并")
                            else:
                                print("t=", t, "-------n=", n, "轨迹在时空上相悖,不可以合并")

                        elif direction == -1:  # 如果是从右向左行驶，那么待匹配目标的终点x坐标应大于第n个目标起点x坐标 且 待匹配目标的终点帧小于第n个目标起点帧  （寻找右侧/上游）
                            # 3.---------------------------------------从右向右，匹配上游-<-<-<---------------------------------------
                            if track_attribute[t][4][0] > track_attribute[n][2][0] and track_attribute[t][5] < track_attribute[n][3]:  # 寻找左侧/上游
                                xn = track_attribute[n][2][0]
                                xt = track_attribute[t][4][0]
                                frame_n = track_attribute[n][3]
                                frame_t = track_attribute[t][5]
                                speed_n = track_attribute[n][1]
                                direction_n = track_attribute[n][0]
                                Similarity, err = position_x_Similarity(xn, xt, frame_n, frame_t, speed_n, direction_n)
                                if Similarity:    # 如果为真，说明第n个目标和第t个目标可以合并
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "<2,可以合并")
                                    if len(merger) == 0:   # 说明还没有值
                                        merger.append([n, t])
                                    else:   # 如果Merger已经存在值了，则进行判断
                                        index_m, flag_m = exist_n_t(merger, n, t)     # index 为当前已存在n或者t的索引号，flag为是否存在n或者t
                                        if flag_m == None:
                                            merger.append([n, t])
                                        elif flag_m == 1:   # 说明只有n存在，则将t添加到n存在的索引号
                                            merger[index_m].append(t)
                                        elif flag_m == -1:  # 说明只有t存在，则将n添加到t存在的索引号
                                            merger[index_m].append(n)
                                        else:   # 剩下就是n和t都存在的情况了，这种情况是，判断t的时候将n一起添加了，然后现在到了n的时候，结果寻找到t了
                                            pass
                                else:
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "≥2,不可以合并")
                            # 4.<-<-<------------------------------------从右向右，匹配下游-------------------------------------------
                            # 待匹配目标的起点x坐标应小于第n个目标终点x坐标 且 待匹配目标的起点帧大于第n个目标终点帧  （寻找左侧/下游）
                            elif track_attribute[t][2][0] < track_attribute[n][4][0] and track_attribute[t][3] > track_attribute[n][5]:  # 寻找左侧/上游
                                xn = track_attribute[n][4][0]
                                xt = track_attribute[t][2][0]
                                frame_n = track_attribute[n][5]
                                frame_t = track_attribute[t][3]
                                speed_n = track_attribute[n][1]
                                direction_n = track_attribute[n][0]
                                Similarity, err = position_x_Similarity(xn, xt, frame_n, frame_t, speed_n, direction_n)
                                if Similarity:  # 如果为真，说明第n个目标和第t个目标可以合并
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "<2,可以合并")
                                    if len(merger) == 0:  # 说明还没有值
                                        merger.append([n, t])
                                    else:  # 如果Merger已经存在值了，则进行判断
                                        index_m, flag_m = exist_n_t(merger, n, t)  # index 为当前已存在n或者t的索引号，flag为是否存在n或者t
                                        if flag_m == None:
                                            merger.append([n, t])
                                        elif flag_m == 1:  # 说明只有n存在，则将t添加到n存在的索引号
                                            merger[index_m].append(t)
                                        elif flag_m == -1:  # 说明只有t存在，则将n添加到t存在的索引号
                                            merger[index_m].append(n)
                                        else:  # 剩下就是n和t都存在的情况了，这种情况是，判断t的时候将n一起添加了，然后现在到了n的时候，结果寻找到t了
                                            pass
                                else:
                                    print("t=", t, "-------n=", n, "，相似度误差:", err, "≥2,不可以合并")
                            else:
                                print("t=", t, "-------n=", n, "轨迹在时空上相悖,不可以合并")
                        else:
                            print("第", n, "个目标出错了，快去检查")
                    else:
                        print("t=", t, "-------n=", n, "t和n的方向不相同，不能进行匹配")
                else:
                    print("t=", t, "-------n=", n, "t和n相等的情况下不需要进行匹配。")

    # ------------------------------------------目标轨迹合并------------------------------------------------
    delete_object = []    # 记录需要删除的目标索引号
    interpolation = []    # 记录需要轨迹插值补全的目标索引号
    min_frame = []        # 记录对应目标索引号的最小帧
    max_frame = []        # 记录对应目标索引号的最大帧
    for m in range(len(merger)):   # 循环每一个合并组合
        interpolation.append(merger[m][0])
        min_frame.append(track_attribute[min(merger[m])][3])
        max_frame.append(track_attribute[max(merger[m])][5])
        for e in range(len(merger[m])):    # 循环每一个目标
            if e != 0:
                delete_object.append(merger[m][e])
                start_frame = track_attribute[merger[m][e]][3]
                end_frame = track_attribute[merger[m][e]][5]
                for frame in range(start_frame, end_frame+1, 1):
                    ALL_object[merger[m][0]][frame] = ALL_object[merger[m][e]][frame]    # 轨迹合并
                    ALL_object_xyzwlh[merger[m][0]][frame] = ALL_object_xyzwlh[merger[m][e]][frame]  # 轨迹合并

    # 对interpolation记录的目标进行插值补全，补全范围为(min_frame, max_frame)
    for inter in range(len(interpolation)):
        xy_incomplete = ALL_object[interpolation[inter]][min_frame[inter]:max_frame[inter]]
        xy_filled = filled_xy(xy_incomplete)
        ALL_object[interpolation[inter]][min_frame[inter]:max_frame[inter]] = xy_filled

        xyzwlh_incomplete = ALL_object_xyzwlh[interpolation[inter]][min_frame[inter]:max_frame[inter]]
        xyzwlh_filled = filled_xyzwlh(xyzwlh_incomplete)
        ALL_object_xyzwlh[interpolation[inter]][min_frame[inter]:max_frame[inter]] = xyzwlh_filled

    # 删除被合并的目标
    print("需要被删除的目标", delete_object)
    diff = 0   # del删除后，ALL_object行号就变了
    for d in range(len(delete_object)):
        del ALL_object[delete_object[d] - diff]   # 删除操作
        del ALL_object_xyzwlh[delete_object[d] - diff]  # 删除操作
        diff += 1

    for new in ALL_object_xyzwlh:
        print(new)
    plot_in_image(ALL_object_xyzwlh)


    # Get the number of frames
    num_frames = len(ALL_object_xyzwlh[0])

    # Loop through each frame
    for frame in range(num_frames):
    # Create a CSV file for the current frame
        with open(f'frame_{frame}.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            # Write the header row
            writer.writerow(['Object Index', 'X', 'Y', 'Z', 'Width', 'Length', 'Height'])

            # Loop through each object in the current frame
            for obj_index, obj_xyzwlh in enumerate(ALL_object_xyzwlh):
                obj_values = obj_xyzwlh[frame]
                if len(obj_values) > 1:
            # Get the x, y, z, w, l, h values for the current object in the current frame
                    obj_values = obj_xyzwlh[frame]
                    
            # Write a row for the current object in the current frame
                    writer.writerow([obj_index, obj_values[0], obj_values[1], obj_values[2],
                                obj_values[3], obj_values[4], obj_values[5]])

                else:
                    writer.writerow([obj_index, 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 'NULL'])


